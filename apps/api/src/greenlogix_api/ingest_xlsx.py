"""openpyxl xlsx ingest with bilingual headers (ORD-01). No geocoding."""

from __future__ import annotations

import logging
from datetime import datetime, time
from io import BytesIO
from openpyxl import load_workbook

from greenlogix_api.models import Order
from greenlogix_api.schemas import ImportErrorItem

log = logging.getLogger("greenlogix")

MAX_UPLOAD_BYTES = 5_000_000
MAX_DATA_ROWS = 500

HEADER_ALIASES: dict[str, frozenset[str]] = {
    "address": frozenset({"address", "địa chỉ"}),
    "lat": frozenset({"lat", "vĩ độ"}),
    "lng": frozenset({"lng", "kinh độ"}),
    "receiver": frozenset({"receiver", "người nhận"}),
    "phone": frozenset({"phone", "sđt"}),
    "kg": frozenset({"kg", "khối lượng"}),
    "window_start": frozenset({"window_start", "giờ bắt đầu"}),
    "window_end": frozenset({"window_end", "giờ kết thúc"}),
    "cargo_type": frozenset({"cargo_type", "loại hàng"}),
    "notes": frozenset({"notes", "ghi chú"}),
}


class IngestLimitError(ValueError):
    pass


def _norm_header(value: object) -> str:
    return str(value or "").strip().lower()


def map_headers(header_row: tuple[object, ...] | list[object]) -> dict[str, int]:
    found: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        key = _norm_header(cell)
        for canon, aliases in HEADER_ALIASES.items():
            if key in aliases:
                found[canon] = idx
                break
    return found


def normalize_window(value: object) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return f"{value.hour:02d}:{value.minute:02d}"
    if isinstance(value, time):
        return f"{value.hour:02d}:{value.minute:02d}"
    if isinstance(value, bool):
        return ""
    if isinstance(value, (int, float)):
        minutes = int(round(float(value) * 24 * 60)) % (24 * 60)
        hours, mins = divmod(minutes, 60)
        return f"{hours:02d}:{mins:02d}"
    text = str(value).strip()
    if "1899" in text and " " in text:
        tail = text.split(" ")[-1]
        return normalize_window(tail)
    if len(text) >= 5 and text[2] == ":":
        return text[:5]
    return text


def _cell(row: tuple[object, ...], index: int | None) -> object:
    if index is None or index >= len(row):
        return None
    return row[index]


def _as_float(value: object) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)  # type: ignore[arg-type]
    except (TypeError, ValueError):
        return None


def parse_xlsx(data: bytes) -> tuple[list[Order], list[ImportErrorItem]]:
    wb = load_workbook(BytesIO(data), data_only=True, read_only=True)
    ws = wb.active
    if ws is None:
        return [], [ImportErrorItem(excel_row=1, field="file", message="empty workbook")]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return [], []
    data_rows = rows[1:]
    if len(data_rows) > MAX_DATA_ROWS:
        raise IngestLimitError("too many rows")
    headers = map_headers(tuple(rows[0] or ()))
    orders: list[Order] = []
    errors: list[ImportErrorItem] = []
    for offset, raw in enumerate(data_rows, start=2):
        row = tuple(raw or ())
        if all(c is None or c == "" for c in row):
            continue
        lat = _as_float(_cell(row, headers.get("lat")))
        lng = _as_float(_cell(row, headers.get("lng")))
        if lat is None:
            errors.append(ImportErrorItem(excel_row=offset, field="lat", message="missing lat"))
            log.info("import skip excel_row=%s field=lat", offset)
            continue
        if lng is None:
            errors.append(ImportErrorItem(excel_row=offset, field="lng", message="missing lng"))
            log.info("import skip excel_row=%s field=lng", offset)
            continue
        kg = _as_float(_cell(row, headers.get("kg"))) or 0.0
        orders.append(
            Order(
                address=str(_cell(row, headers.get("address")) or ""),
                lat=lat,
                lng=lng,
                receiver=str(_cell(row, headers.get("receiver")) or ""),
                phone=str(_cell(row, headers.get("phone")) or ""),
                kg=kg,
                window_start=normalize_window(_cell(row, headers.get("window_start"))),
                window_end=normalize_window(_cell(row, headers.get("window_end"))),
                cargo_type=str(_cell(row, headers.get("cargo_type")) or "thuong"),
                notes=str(_cell(row, headers.get("notes")) or ""),
                excel_row=offset,
                status="pending",
            )
        )
        log.info("import excel_row=%s", offset)
    return orders, errors
