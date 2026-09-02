"""GET /report.xlsx openpyxl workbook matches GET /report JSON."""

from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlmodel import Session

from greenlogix_api import db as dbmod
from greenlogix_api.models import Order, Vehicle

AUTH = {"Authorization": "Bearer DEMO"}
XLSX_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _tiny_zigzag() -> None:
    with Session(dbmod.engine) as session:
        session.add(
            Vehicle(
                plate="51C-ZIG.01",
                type="van",
                capacity_kg=2000,
                fuel="petrol",
                l_per_100km=10.0,
                status="ready",
            )
        )
        session.add(
            Order(
                address="North close A",
                lat=10.81,
                lng=106.661,
                kg=10,
                excel_row=2,
                window_start="08:00",
                window_end="12:00",
            )
        )
        session.add(
            Order(
                address="Far south-east B",
                lat=10.70,
                lng=106.80,
                kg=10,
                excel_row=3,
                window_start="08:00",
                window_end="12:00",
            )
        )
        session.add(
            Order(
                address="North close C",
                lat=10.82,
                lng=106.661,
                kg=10,
                excel_row=4,
                window_start="08:00",
                window_end="12:00",
            )
        )
        session.commit()


def test_report_xlsx_matches_json_and_spreadsheetml(demo_client) -> None:
    _tiny_zigzag()
    optimized = demo_client.post(
        "/optimize",
        headers=AUTH,
        json={"cluster_radius_km": 3.0},
    )
    assert optimized.status_code == 200
    json_body = demo_client.get("/report", headers=AUTH)
    assert json_body.status_code == 200
    payload = json_body.json()
    xlsx = demo_client.get("/report.xlsx", headers=AUTH)
    assert xlsx.status_code == 200
    assert "spreadsheetml" in (xlsx.headers.get("content-type") or "")
    assert xlsx.headers["content-type"].startswith(XLSX_TYPE.split(";")[0])
    wb = load_workbook(BytesIO(xlsx.content))
    ws = wb.active
    assert ws is not None
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    assert headers == ["metric", "baseline", "optimized", "delta", "delta_pct"]
    rows = {row[0]: row for row in ws.iter_rows(min_row=2, values_only=True)}
    for metric, pct_key in (
        ("km", "km_pct"),
        ("litres", "litres_pct"),
        ("kg_co2", "kg_co2_pct"),
    ):
        assert metric in rows
        _name, baseline, optimized_v, delta, delta_pct = rows[metric]
        assert baseline == pytest.approx(payload["baseline"][metric])
        assert optimized_v == pytest.approx(payload["optimized"][metric])
        assert delta == pytest.approx(payload["delta"][metric])
        assert delta_pct == pytest.approx(payload["delta"][pct_key])


def test_report_xlsx_requires_bearer(demo_client) -> None:
    assert demo_client.get("/report.xlsx").status_code == 401
